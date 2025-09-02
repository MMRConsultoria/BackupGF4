# pages/PainelResultados.py
import streamlit as st
st.set_page_config(page_title="Vendas Diarias", layout="wide")

# ===== Imports =====
import pandas as pd
from io import BytesIO
from datetime import datetime
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import json
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImg
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
import pytz
import io

# üîí Bloqueio de acesso
if not st.session_state.get("acesso_liberado"):
    st.stop()

# ===== CSS (visual) =====
st.markdown("""
    <style>
        [data-testid="stToolbar"] { visibility: hidden; height: 0%; position: fixed; }
        .stSpinner { visibility: visible !important; }
        .stApp { background-color: #f9f9f9; }
        div[data-baseweb="tab-list"] { margin-top: 20px; }
        button[data-baseweb="tab"] {
            background-color: #f0f2f6; border-radius: 10px;
            padding: 10px 20px; margin-right: 10px;
            transition: all 0.3s ease; font-size: 16px; font-weight: 600;
        }
        button[data-baseweb="tab"]:hover { background-color: #dce0ea; color: black; }
        button[data-baseweb="tab"][aria-selected="true"] { background-color: #0366d6; color: white; }

        /* multiselect sem tags coloridas */
        div[data-testid="stMultiSelect"] [data-baseweb="tag"] { background-color: transparent !important; border: none !important; color: black !important; }
        div[data-testid="stMultiSelect"] [data-baseweb="tag"] * { color: black !important; fill: black !important; }
        div[data-testid="stMultiSelect"] > div { background-color: transparent !important; }
    </style>
""", unsafe_allow_html=True)

# ===== Cabe√ßalho =====
st.markdown("""
    <div style='display: flex; align-items: center; gap: 10px; margin-bottom: 12px;'>
        <img src='https://img.icons8.com/color/48/graph.png' width='40'/>
        <h1 style='display: inline; margin: 0; font-size: 2.0rem;'>Rateio</h1>
    </div>
""", unsafe_allow_html=True)

with st.spinner("‚è≥ Processando..."):
    # ===== Conex√£o + dados base (comum √†s abas) =====
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials_dict = json.loads(st.secrets["GOOGLE_SERVICE_ACCOUNT"])
    credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
    gc = gspread.authorize(credentials)
    planilha = gc.open("Vendas diarias")

    df_empresa = pd.DataFrame(planilha.worksheet("Tabela Empresa").get_all_records())
    df_vendas  = pd.DataFrame(planilha.worksheet("Fat Sistema Externo").get_all_records())

    # Normaliza√ß√£o comum
    df_empresa.columns = df_empresa.columns.str.strip()
    df_vendas.columns  = df_vendas.columns.str.strip()
    if "Loja" in df_empresa.columns:
        df_empresa["Loja"] = df_empresa["Loja"].astype(str).str.strip().str.upper()
    if "Grupo" in df_empresa.columns:
        df_empresa["Grupo"] = df_empresa["Grupo"].astype(str).str.strip()
    if "Data" in df_vendas.columns:
        df_vendas["Data"] = pd.to_datetime(df_vendas["Data"], dayfirst=True, errors="coerce")
    if "Loja" in df_vendas.columns:
        df_vendas["Loja"] = df_vendas["Loja"].astype(str).str.strip().str.upper()
    if "Grupo" in df_vendas.columns:
        df_vendas["Grupo"] = df_vendas["Grupo"].astype(str).str.strip()
    # Merge com Tipo
    if {"Loja","Tipo"}.issubset(df_empresa.columns):
        df_vendas = df_vendas.merge(df_empresa[["Loja","Tipo"]], on="Loja", how="left")
    else:
        df_vendas["Tipo"] = df_vendas.get("Tipo", "")

    # Converter Fat.Total para n√∫mero
    if "Fat.Total" in df_vendas.columns:
        df_vendas["Fat.Total"] = (
            df_vendas["Fat.Total"].astype(str)
            .str.replace("R$", "", regex=False)
            .str.replace("(", "-", regex=False)
            .str.replace(")", "", regex=False)
            .str.replace(" ", "", regex=False)
            .str.replace(".", "", regex=False)
            .str.replace(",", ".", regex=False)
        )
        df_vendas["Fat.Total"] = pd.to_numeric(df_vendas["Fat.Total"], errors="coerce").fillna(0.0)
    else:
        df_vendas["Fat.Total"] = 0.0

    # ===== Abas =====
    aba1, aba2 = st.tabs(["üìÑ %Faturamento", "üîÑ Volumetria"])

    # ----------------------------------------------------------------------
    # ABA 1 - % FATURAMENTO  (tudo desta aba fica aqui dentro)
    # ----------------------------------------------------------------------
    with aba1:
        # --------- Filtros ---------
        col1, col2, col3 = st.columns([1, 1, 2])

        with col1:
            tipos = sorted(df_vendas["Tipo"].dropna().unique())
            tipos.insert(0, "Todos")
            tipo_sel = st.selectbox("üè™ Tipo:", options=tipos, index=0, key="tipo_fat")

        with col2:
            grupos = sorted(df_vendas["Grupo"].dropna().unique())
            grupos.insert(0, "Todos")
            grupo_sel = st.selectbox("üë• Grupo:", options=grupos, index=0, key="grupo_fat")

        with col3:
            df_vendas["Mes/Ano"] = df_vendas["Data"].dt.strftime("%m/%Y")

            def _ord_key(mmyyyy: str):
                try:
                    return datetime.strptime("01/" + str(mmyyyy), "%d/%m/%Y")
                except Exception:
                    return datetime.min

            meses_opts = sorted([m for m in df_vendas["Mes/Ano"].dropna().unique()], key=_ord_key)
            mes_atual = datetime.today().strftime("%m/%Y")
            default_meses = [mes_atual] if meses_opts and mes_atual in meses_opts else (meses_opts[-1:] if meses_opts else [])

            if meses_opts:
                meses_sel = st.multiselect(
                    "üóìÔ∏è Selecione os meses:",
                    options=meses_opts,
                    default=default_meses,
                    key="ms_meses_fat"
                )
            else:
                st.warning("‚ö†Ô∏è Nenhum m√™s dispon√≠vel nos dados (verifique a coluna 'Data').")
                meses_sel = []

        # --------- Aplica filtros ---------
        if meses_sel:
            df_f = df_vendas[df_vendas["Mes/Ano"].isin(meses_sel)].copy()
        else:
            df_f = df_vendas.iloc[0:0].copy()

        df_f["Per√≠odo"] = df_f["Data"].dt.strftime("%m/%Y")
        if tipo_sel != "Todos":
            df_f = df_f[df_f["Tipo"] == tipo_sel]
        if grupo_sel != "Todos":
            df_f = df_f[df_f["Grupo"] == grupo_sel]

        # --------- Agrupamento ---------
        metric = "Fat.Total"  # m√©trica desta aba
        if grupo_sel == "Todos":
            chaves = ["Tipo", "Grupo"]
        else:
            chaves = ["Grupo", "Loja"]

        df_ag = df_f.groupby(chaves + ["Per√≠odo"], as_index=False)[metric].sum()
        df_fin = df_ag.groupby(chaves, as_index=False)[metric].sum().rename(columns={metric: "Total"})
        df_fin["Rateio"] = 0.0

        # --------- % e Subtotais ---------
        if grupo_sel == "Todos":
            total_geral = df_fin["Total"].sum()
            df_fin["% Total"] = (df_fin["Total"] / total_geral) if total_geral else 0.0

            subt = df_fin.groupby("Tipo")["Total"].sum().reset_index().sort_values(by="Total", ascending=False)
            ordem_tipos = subt["Tipo"].tolist()
            df_fin["ord_tipo"] = df_fin["Tipo"].apply(lambda x: ordem_tipos.index(x) if x in ordem_tipos else 999)
            df_fin = df_fin.sort_values(by=["ord_tipo", "Total"], ascending=[True, False]).drop(columns="ord_tipo")

            linhas = []
            for t in ordem_tipos:
                bloco = df_fin[df_fin["Tipo"] == t].copy()
                linhas.append(bloco)
                subtotal = bloco.drop(columns=["Tipo", "Grupo"]).sum(numeric_only=True)
                subtotal["Tipo"] = t
                subtotal["Grupo"] = f"Subtotal {t}"
                linhas.append(pd.DataFrame([subtotal]))
            df_fin = pd.concat(linhas, ignore_index=True)
        else:
            total_geral = df_fin["Total"].sum()
            df_fin["% Total"] = (df_fin["Total"] / total_geral) if total_geral else 0.0

            df_fin = df_fin.sort_values(by=["Grupo", "Total"], ascending=[True, False])

            linhas = []
            for g in df_fin["Grupo"].unique():
                bloco = df_fin[df_fin["Grupo"] == g].copy()
                linhas.append(bloco)
                subtotal = bloco.drop(columns=["Grupo", "Loja"]).sum(numeric_only=True)
                subtotal["Grupo"] = g
                subtotal["Loja"] = f"Subtotal {g}"
                linhas.append(pd.DataFrame([subtotal]))
            df_fin = pd.concat(linhas, ignore_index=True)

        # --------- TOTAL no topo ---------
        cols_drop = [c for c in ["Tipo","Grupo","Loja"] if c in df_fin.columns]
        apenas = df_fin.copy()
        for c in cols_drop:
            apenas = apenas[~apenas[c].astype(str).str.startswith("Subtotal", na=False)]
        linha_total = apenas.drop(columns=cols_drop, errors="ignore").sum(numeric_only=True)
        for c in cols_drop:
            linha_total[c] = ""
        linha_total[cols_drop[0] if cols_drop else "Grupo"] = "TOTAL"
        df_fin = pd.concat([pd.DataFrame([linha_total]), df_fin], ignore_index=True)

        # --------- RATEIO ---------
        df_fin["% Total"] = 0.0
        df_fin["Rateio"] = 0.0

        if grupo_sel == "Todos":
            def moeda_para_float(s: str) -> float:
                try: return float(s.replace(".", "").replace(",", "."))
                except: return 0.0

            tipos_unicos = [t for t in df_fin["Tipo"].dropna().unique()
                            if str(t).strip() not in ["", "TOTAL"] and not str(t).startswith("Subtotal")]
            valores_rateio = {}
            COLS_POR_LINHA = 3
            for i in range(0, len(tipos_unicos), COLS_POR_LINHA):
                linha = tipos_unicos[i:i+COLS_POR_LINHA]
                cols = st.columns(len(linha))
                for c, t in zip(cols, linha):
                    with c:
                        valor_str = st.text_input(f"üí∞ Rateio ‚Äî {t}", value="0,00", key=f"rateio_{t}_fat")
                        valores_rateio[t] = moeda_para_float(valor_str)

            for t in df_fin["Tipo"].unique():
                mask = ((df_fin["Tipo"] == t) &
                        (~df_fin["Grupo"].astype(str).str.startswith("Subtotal")) &
                        (df_fin["Grupo"] != "TOTAL"))
                subtotal_t = df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "Total"].sum()
                if subtotal_t > 0:
                    df_fin.loc[mask, "% Total"] = (df_fin.loc[mask, "Total"] / subtotal_t) * 100
                df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "% Total"] = 100

                valor_rateio = valores_rateio.get(t, 0.0)
                df_fin.loc[mask, "Rateio"] = df_fin.loc[mask, "% Total"] / 100 * valor_rateio
                df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "Rateio"] = df_fin.loc[mask, "Rateio"].sum()
        else:
            total_rateio = st.number_input(f"üí∞ Rateio ‚Äî {grupo_sel}",
                                           min_value=0.0, step=100.0, format="%.2f",
                                           key=f"rateio_{grupo_sel}_fat")
            mask_lojas = ((df_fin["Grupo"] == grupo_sel) &
                          (~df_fin["Loja"].astype(str).str.startswith("Subtotal")) &
                          (df_fin["Loja"] != "TOTAL"))
            subtotal_g = df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "Total"].sum()
            if subtotal_g > 0:
                df_fin.loc[mask_lojas, "% Total"] = (df_fin.loc[mask_lojas, "Total"] / subtotal_g) * 100
                df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "% Total"] = 100
                df_fin.loc[mask_lojas, "Rateio"] = df_fin.loc[mask_lojas, "% Total"] / 100 * total_rateio
                df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "Rateio"] = df_fin.loc[mask_lojas, "Rateio"].sum()
        # === Reordenar colunas (Aba 1) ===
        if grupo_sel == "Todos":
            col_order = ["Tipo", "Grupo", "Total", "% Total", "Rateio"]
        else:
            # quando filtra um grupo espec√≠fico, aparece "Loja"
            col_order = ["Grupo", "Loja", "Total", "% Total", "Rateio"]
        
        # mant√©m s√≥ as colunas nessa ordem (as que existirem)
        df_fin = df_fin[[c for c in col_order if c in df_fin.columns]]
        # --------- Visual ---------
        df_view = df_fin.copy()
        def fmt_moeda(v):
            try: return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except: return v
        for c in ["Total","Rateio"]:
            if c in df_view.columns:
                df_view[c] = df_view[c].apply(lambda x: fmt_moeda(x) if pd.notnull(x) and x != "" else x)
        if "% Total" in df_view.columns:
            df_view["% Total"] = pd.to_numeric(df_view["% Total"], errors="coerce").apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else "")

        def aplicar_estilo_fat(df_in):
            def estilo(row):
                if "Grupo" in df_in.columns and row["Grupo"] == "TOTAL":
                    return ["background-color: #f4b084; font-weight: bold"] * len(row)
                if "Loja" in df_in.columns and isinstance(row.get("Loja",""), str) and row["Loja"].startswith("Subtotal"):
                    return ["background-color: #d9d9d9; font-weight: bold"] * len(row)
                if "Grupo" in df_in.columns and isinstance(row.get("Grupo",""), str) and row["Grupo"].startswith("Subtotal"):
                    return ["background-color: #d9d9d9; font-weight: bold"] * len(row)
                return ["" for _ in row]
            return df_in.style.apply(estilo, axis=1)

        st.dataframe(aplicar_estilo_fat(df_view), use_container_width=True, height=700)

        # --------- Exportar Excel ---------
        df_excel = df_fin.copy()
        if "% Total" in df_excel.columns:
            df_excel["% Total"] = pd.to_numeric(df_excel["% Total"], errors="coerce") / 100
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df_excel.to_excel(writer, index=False, sheet_name="Relat√≥rio")
        out.seek(0)
        wb = load_workbook(out); ws = wb["Relat√≥rio"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for cell in ws[1]:
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center_alignment; cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            grupo_val = None
            try: grupo_val = row[1].value
            except: pass
            estilo_fundo = None
            if isinstance(grupo_val, str):
                if grupo_val.strip().upper() == "TOTAL": estilo_fundo = PatternFill("solid", fgColor="F4B084")
                elif "SUBTOTAL" in grupo_val.strip().upper(): estilo_fundo = PatternFill("solid", fgColor="D9D9D9")
            for cell in row:
                cell.border = border; cell.alignment = center_alignment
                if estilo_fundo: cell.fill = estilo_fundo
                col_name = ws.cell(row=1, column=cell.column).value
                if isinstance(cell.value, (int,float)):
                    cell.number_format = '0.00%' if col_name == "% Total" else '"R$" #,##0.00'

        for i, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        for col_nome in ["Tipo","Grupo","Loja"]:
            if col_nome in df_excel.columns:
                col_idx = df_excel.columns.get_loc(col_nome) + 1
                for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for c in cell: c.alignment = Alignment(horizontal="left")

        out_final = BytesIO(); wb.save(out_final); out_final.seek(0)
        st.download_button("üì• Baixar Excel", data=out_final,
                           file_name="Resumo_%Faturamento.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_excel_fat")

        # --------- Exportar PDF ---------
        usuario = st.session_state.get("usuario_logado", "Usu√°rio Desconhecido")
        sele = meses_sel
        if not sele: mes_rateio = "(sem dados)"
        elif len(sele) == 1: mes_rateio = sele[0]
        elif len(sele) == 2: mes_rateio = f"{sele[0]} e {sele[1]}"
        else: mes_rateio = f"{', '.join(sele[:-1])} e {sele[-1]}"

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=20, rightMargin=20)
        elems = []; estilos = getSampleStyleSheet(); normal = estilos["Normal"]; h1 = estilos["Heading1"]
        try:
            logo_url = "https://raw.githubusercontent.com/MMRConsultoria/mmr-site/main/logo_grupofit.png"
            img = RLImg(logo_url, width=100, height=40); elems.append(img)
        except: pass
        elems.append(Paragraph(f"<b>Rateio - {mes_rateio}</b>", h1))
        fuso = pytz.timezone("America/Sao_Paulo")
        data_ger = datetime.now(fuso).strftime("%d/%m/%Y %H:%M")
        elems.append(Paragraph(f"<b>Usu√°rio:</b> {usuario}", normal))
        elems.append(Paragraph(f"<b>Data de Gera√ß√£o:</b> {data_ger}", normal)); elems.append(Spacer(1,12))
        dados = [df_view.columns.tolist()] + df_view.values.tolist()
        tabela = Table(dados, repeatRows=1)
        tabela.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#003366")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("ALIGN",(1,1),(-1,-1),"CENTER"),
            ("ALIGN",(0,0),(0,-1),"LEFT"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("BOTTOMPADDING",(0,0),(-1,0),8),
            ("GRID",(0,0),(-1,-1),0.25,colors.grey),
        ]))
        for i in range(1, len(dados)):
            txt = str(dados[i][1]).strip().lower() if len(dados[i])>1 else ""
            if "subtotal" in txt or txt == "total":
                tabela.setStyle(TableStyle([("BACKGROUND",(0,i),(-1,i),colors.HexColor("#BFBFBF")),
                                            ("FONTNAME",(0,i),(-1,i),"Helvetica-Bold")]))
            else:
                tabela.setStyle(TableStyle([("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F2F2F2"))]))
        elems.append(tabela); doc.build(elems)
        pdf_bytes = buf.getvalue(); buf.close()
        st.download_button("üìÑ Baixar PDF", data=pdf_bytes,
                           file_name=f"Rateio_%Faturamento_{datetime.now().strftime('%Y%m%d')}.pdf",
                           mime="application/pdf", key="dl_pdf_fat")

    # ----------------------------------------------------------------------
    # ABA 2 - VOLUMETRIA  (tudo desta aba fica aqui dentro)
    # ----------------------------------------------------------------------
    with aba2:
        # üëâ quando voc√™ me disser qual coluna √© a de quantidade (ex.: 'Qtde'),
        #    troco 'metric_vol' abaixo para essa coluna.
        metric_vol = "Fat.Total"  # TEMPOR√ÅRIO: mesma m√©trica at√© definirmos a coluna de volume

        # --------- Filtros ---------
        col1, col2, col3 = st.columns([1, 1, 2])

        with col1:
            tipos = sorted(df_vendas["Tipo"].dropna().unique())
            tipos.insert(0, "Todos")
            tipo_sel = st.selectbox("üè™ Tipo:", options=tipos, index=0, key="tipo_vol")

        with col2:
            grupos = sorted(df_vendas["Grupo"].dropna().unique())
            grupos.insert(0, "Todos")
            grupo_sel = st.selectbox("üë• Grupo:", options=grupos, index=0, key="grupo_vol")

        with col3:
            df_vendas["Mes/Ano"] = df_vendas["Data"].dt.strftime("%m/%Y")

            def _ord_key2(mmyyyy: str):
                try:
                    return datetime.strptime("01/" + str(mmyyyy), "%d/%m/%Y")
                except Exception:
                    return datetime.min

            meses_opts = sorted([m for m in df_vendas["Mes/Ano"].dropna().unique()], key=_ord_key2)
            mes_atual = datetime.today().strftime("%m/%Y")
            default_meses = [mes_atual] if meses_opts and mes_atual in meses_opts else (meses_opts[-1:] if meses_opts else [])

            if meses_opts:
                meses_sel = st.multiselect(
                    "üóìÔ∏è Selecione os meses:",
                    options=meses_opts,
                    default=default_meses,
                    key="ms_meses_vol"
                )
            else:
                st.warning("‚ö†Ô∏è Nenhum m√™s dispon√≠vel nos dados (verifique a coluna 'Data').")
                meses_sel = []

        # --------- Aplica filtros ---------
        if meses_sel:
            df_f = df_vendas[df_vendas["Mes/Ano"].isin(meses_sel)].copy()
        else:
            df_f = df_vendas.iloc[0:0].copy()

        df_f["Per√≠odo"] = df_f["Data"].dt.strftime("%m/%Y")
        if tipo_sel != "Todos":
            df_f = df_f[df_f["Tipo"] == tipo_sel]
        if grupo_sel != "Todos":
            df_f = df_f[df_f["Grupo"] == grupo_sel]

        # --------- Agrupamento ---------
        if grupo_sel == "Todos":
            chaves = ["Tipo", "Grupo"]
        else:
            chaves = ["Grupo", "Loja"]

        df_ag = df_f.groupby(chaves + ["Per√≠odo"], as_index=False)[metric_vol].sum()
        df_fin = df_ag.groupby(chaves, as_index=False)[metric_vol].sum().rename(columns={metric_vol: "Total"})
        df_fin["Rateio"] = 0.0

        # --------- % e Subtotais ---------
        if grupo_sel == "Todos":
            total_geral = df_fin["Total"].sum()
            df_fin["% Total"] = (df_fin["Total"] / total_geral) if total_geral else 0.0

            subt = df_fin.groupby("Tipo")["Total"].sum().reset_index().sort_values(by="Total", ascending=False)
            ordem_tipos = subt["Tipo"].tolist()
            df_fin["ord_tipo"] = df_fin["Tipo"].apply(lambda x: ordem_tipos.index(x) if x in ordem_tipos else 999)
            df_fin = df_fin.sort_values(by=["ord_tipo", "Total"], ascending=[True, False]).drop(columns="ord_tipo")

            linhas = []
            for t in ordem_tipos:
                bloco = df_fin[df_fin["Tipo"] == t].copy()
                linhas.append(bloco)
                subtotal = bloco.drop(columns=["Tipo", "Grupo"]).sum(numeric_only=True)
                subtotal["Tipo"] = t
                subtotal["Grupo"] = f"Subtotal {t}"
                linhas.append(pd.DataFrame([subtotal]))
            df_fin = pd.concat(linhas, ignore_index=True)
        else:
            total_geral = df_fin["Total"].sum()
            df_fin["% Total"] = (df_fin["Total"] / total_geral) if total_geral else 0.0

            df_fin = df_fin.sort_values(by=["Grupo", "Total"], ascending=[True, False])

            linhas = []
            for g in df_fin["Grupo"].unique():
                bloco = df_fin[df_fin["Grupo"] == g].copy()
                linhas.append(bloco)
                subtotal = bloco.drop(columns=["Grupo", "Loja"]).sum(numeric_only=True)
                subtotal["Grupo"] = g
                subtotal["Loja"] = f"Subtotal {g}"
                linhas.append(pd.DataFrame([subtotal]))
            df_fin = pd.concat(linhas, ignore_index=True)

        # --------- TOTAL no topo ---------
        cols_drop = [c for c in ["Tipo","Grupo","Loja"] if c in df_fin.columns]
        apenas = df_fin.copy()
        for c in cols_drop:
            apenas = apenas[~apenas[c].astype(str).str.startswith("Subtotal", na=False)]
        linha_total = apenas.drop(columns=cols_drop, errors="ignore").sum(numeric_only=True)
        for c in cols_drop:
            linha_total[c] = ""
        linha_total[cols_drop[0] if cols_drop else "Grupo"] = "TOTAL"
        df_fin = pd.concat([pd.DataFrame([linha_total]), df_fin], ignore_index=True)

        # --------- RATEIO ---------
        df_fin["% Total"] = 0.0
        df_fin["Rateio"] = 0.0

        if grupo_sel == "Todos":
            def moeda_para_float2(s: str) -> float:
                try: return float(s.replace(".", "").replace(",", "."))
                except: return 0.0
            tipos_unicos = [t for t in df_fin["Tipo"].dropna().unique()
                            if str(t).strip() not in ["", "TOTAL"] and not str(t).startswith("Subtotal")]
            valores_rateio = {}
            COLS_POR_LINHA = 3
            for i in range(0, len(tipos_unicos), COLS_POR_LINHA):
                linha = tipos_unicos[i:i+COLS_POR_LINHA]
                cols = st.columns(len(linha))
                for c, t in zip(cols, linha):
                    with c:
                        valor_str = st.text_input(f"üí∞ Rateio ‚Äî {t}", value="0,00", key=f"rateio_{t}_vol")
                        valores_rateio[t] = moeda_para_float2(valor_str)

            for t in df_fin["Tipo"].unique():
                mask = ((df_fin["Tipo"] == t) &
                        (~df_fin["Grupo"].astype(str).str.startswith("Subtotal")) &
                        (df_fin["Grupo"] != "TOTAL"))
                subtotal_t = df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "Total"].sum()
                if subtotal_t > 0:
                    df_fin.loc[mask, "% Total"] = (df_fin.loc[mask, "Total"] / subtotal_t) * 100
                df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "% Total"] = 100

                valor_rateio = valores_rateio.get(t, 0.0)
                df_fin.loc[mask, "Rateio"] = df_fin.loc[mask, "% Total"] / 100 * valor_rateio
                df_fin.loc[df_fin["Grupo"] == f"Subtotal {t}", "Rateio"] = df_fin.loc[mask, "Rateio"].sum()
        else:
            total_rateio = st.number_input(f"üí∞ Rateio ‚Äî {grupo_sel}",
                                           min_value=0.0, step=100.0, format="%.2f",
                                           key=f"rateio_{grupo_sel}_vol")
            mask_lojas = ((df_fin["Grupo"] == grupo_sel) &
                          (~df_fin["Loja"].astype(str).str.startswith("Subtotal")) &
                          (df_fin["Loja"] != "TOTAL"))
            subtotal_g = df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "Total"].sum()
            if subtotal_g > 0:
                df_fin.loc[mask_lojas, "% Total"] = (df_fin.loc[mask_lojas, "Total"] / subtotal_g) * 100
                df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "% Total"] = 100
                df_fin.loc[mask_lojas, "Rateio"] = df_fin.loc[mask_lojas, "% Total"] / 100 * total_rateio
                df_fin.loc[df_fin["Loja"] == f"Subtotal {grupo_sel}", "Rateio"] = df_fin.loc[mask_lojas, "Rateio"].sum()
        # === Reordenar colunas (Aba 2) ===
        if grupo_sel == "Todos":
            col_order = ["Tipo", "Grupo", "Total", "% Total", "Rateio"]
        else:
            col_order = ["Grupo", "Loja", "Total", "% Total", "Rateio"]
        
        df_fin = df_fin[[c for c in col_order if c in df_fin.columns]]

        # --------- Visual ---------
        df_view = df_fin.copy()
        def fmt_moeda2(v):
            try: return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except: return v
        for c in ["Total","Rateio"]:
            if c in df_view.columns:
                df_view[c] = df_view[c].apply(lambda x: fmt_moeda2(x) if pd.notnull(x) and x != "" else x)
        if "% Total" in df_view.columns:
            df_view["% Total"] = pd.to_numeric(df_view["% Total"], errors="coerce").apply(lambda x: f"{x:.2f}%" if pd.notnull(x) else "")

        def aplicar_estilo_vol(df_in):
            def estilo(row):
                if "Grupo" in df_in.columns and row["Grupo"] == "TOTAL":
                    return ["background-color: #f4b084; font-weight: bold"] * len(row)
                if "Loja" in df_in.columns and isinstance(row.get("Loja",""), str) and row["Loja"].startswith("Subtotal"):
                    return ["background-color: #d9d9d9; font-weight: bold"] * len(row)
                if "Grupo" in df_in.columns and isinstance(row.get("Grupo",""), str) and row["Grupo"].startswith("Subtotal"):
                    return ["background-color: #d9d9d9; font-weight: bold"] * len(row)
                return ["" for _ in row]
            return df_in.style.apply(estilo, axis=1)

        st.dataframe(aplicar_estilo_vol(df_view), use_container_width=True, height=700)

        # --------- Exportar Excel ---------
        df_excel = df_fin.copy()
        if "% Total" in df_excel.columns:
            df_excel["% Total"] = pd.to_numeric(df_excel["% Total"], errors="coerce") / 100
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as writer:
            df_excel.to_excel(writer, index=False, sheet_name="Relat√≥rio")
        out.seek(0)
        wb = load_workbook(out); ws = wb["Relat√≥rio"]

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="305496")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

        for cell in ws[1]:
            cell.font = header_font; cell.fill = header_fill
            cell.alignment = center_alignment; cell.border = border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
            grupo_val = None
            try: grupo_val = row[1].value
            except: pass
            estilo_fundo = None
            if isinstance(grupo_val, str):
                if grupo_val.strip().upper() == "TOTAL": estilo_fundo = PatternFill("solid", fgColor="F4B084")
                elif "SUBTOTAL" in grupo_val.strip().upper(): estilo_fundo = PatternFill("solid", fgColor="D9D9D9")
            for cell in row:
                cell.border = border; cell.alignment = center_alignment
                if estilo_fundo: cell.fill = estilo_fundo
                col_name = ws.cell(row=1, column=cell.column).value
                if isinstance(cell.value, (int,float)):
                    cell.number_format = '0.00%' if col_name == "% Total" else '"R$" #,##0.00'

        for i, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=ws.max_row), start=1):
            max_len = max((len(str(c.value)) for c in col_cells if c.value), default=0)
            ws.column_dimensions[get_column_letter(i)].width = max_len + 2

        for col_nome in ["Tipo","Grupo","Loja"]:
            if col_nome in df_excel.columns:
                col_idx = df_excel.columns.get_loc(col_nome) + 1
                for cell in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for c in cell: c.alignment = Alignment(horizontal="left")

        out_final = BytesIO(); wb.save(out_final); out_final.seek(0)
        st.download_button("üì• Baixar Excel", data=out_final,
                           file_name="Resumo_Volumetria.xlsx",
                           mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           key="dl_excel_vol")

        # --------- Exportar PDF ---------
        usuario = st.session_state.get("usuario_logado", "Usu√°rio Desconhecido")
        sele = meses_sel
        if not sele: mes_rateio = "(sem dados)"
        elif len(sele) == 1: mes_rateio = sele[0]
        elif len(sele) == 2: mes_rateio = f"{sele[0]} e {sele[1]}"
        else: mes_rateio = f"{', '.join(sele[:-1])} e {sele[-1]}"

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=30, bottomMargin=30, leftMargin=20, rightMargin=20)
        elems = []; estilos = getSampleStyleSheet(); normal = estilos["Normal"]; h1 = estilos["Heading1"]
        try:
            logo_url = "https://raw.githubusercontent.com/MMRConsultoria/mmr-site/main/logo_grupofit.png"
            img = RLImg(logo_url, width=100, height=40); elems.append(img)
        except: pass
        elems.append(Paragraph(f"<b>Rateio - {mes_rateio}</b>", h1))
        fuso = pytz.timezone("America/Sao_Paulo")
        data_ger = datetime.now(fuso).strftime("%d/%m/%Y %H:%M")
        elems.append(Paragraph(f"<b>Usu√°rio:</b> {usuario}", normal))
        elems.append(Paragraph(f"<b>Data de Gera√ß√£o:</b> {data_ger}", normal)); elems.append(Spacer(1,12))
        dados = [df_view.columns.tolist()] + df_view.values.tolist()
        tabela = Table(dados, repeatRows=1)
        tabela.setStyle(TableStyle([
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#003366")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("ALIGN",(1,1),(-1,-1),"CENTER"),
            ("ALIGN",(0,0),(0,-1),"LEFT"),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("BOTTOMPADDING",(0,0),(-1,0),8),
            ("GRID",(0,0),(-1,-1),0.25,colors.grey),
        ]))
        for i in range(1, len(dados)):
            txt = str(dados[i][1]).strip().lower() if len(dados[i])>1 else ""
            if "subtotal" in txt or txt == "total":
                tabela.setStyle(TableStyle([("BACKGROUND",(0,i),(-1,i),colors.HexColor("#BFBFBF")),
                                            ("FONTNAME",(0,i),(-1,i),"Helvetica-Bold")]))
            else:
                tabela.setStyle(TableStyle([("BACKGROUND",(0,i),(-1,i),colors.HexColor("#F2F2F2"))]))
        elems.append(tabela); doc.build(elems)
        pdf_bytes = buf.getvalue(); buf.close()
        st.download_button("üìÑ Baixar PDF", data=pdf_bytes,
                           file_name=f"Rateio_Volumetria_{datetime.now().strftime('%Y%m%d')}.pdf",
                           mime="application/pdf", key="dl_pdf_vol")
