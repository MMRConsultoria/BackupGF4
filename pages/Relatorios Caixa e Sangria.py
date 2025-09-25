# pages/05_Relatorios_Caixa_Sangria.py
import streamlit as st


import pandas as pd
import numpy as np
import re, json
from io import BytesIO
import gspread
from oauth2client.service_account import ServiceAccountCredentials

# Bloqueio opcional (mantenha se voc√™ j√° usa login/sess√£o)
if not st.session_state.get("acesso_liberado"):
    st.stop()

# ======================
# CSS (opcional)
# ======================
st.markdown("""
<style>
[data-testid="stToolbar"]{visibility:hidden;height:0;position:fixed}
</style>
""", unsafe_allow_html=True)

with st.spinner("‚è≥ Carregando dados..."):
    # ============ Conex√£o com Google Sheets ============
    scope = ["https://spreadsheets.google.com/feeds", "https://www.googleapis.com/auth/drive"]
    credentials_dict = json.loads(st.secrets["GOOGLE_SERVICE_ACCOUNT"])
    credentials = ServiceAccountCredentials.from_json_keyfile_dict(credentials_dict, scope)
    gc = gspread.authorize(credentials)
    planilha_empresa = gc.open("Vendas diarias")

    # Tabela Empresa (para mapear Grupo/Loja/Tipo/PDV etc.)
    df_empresa = pd.DataFrame(planilha_empresa.worksheet("Tabela Empresa").get_all_records())
    df_empresa.columns = [str(c).strip() for c in df_empresa.columns]
    # ================================
    # 2. Configura√ß√£o inicial do app
    # ================================
    
    
    # üé® Estilizar abas
    st.markdown("""
        <style>
        .stApp { background-color: #f9f9f9; }
        div[data-baseweb="tab-list"] { margin-top: 20px; }
        button[data-baseweb="tab"] {
            background-color: #f0f2f6;
            border-radius: 10px;
            padding: 10px 20px;
            margin-right: 10px;
            transition: all 0.3s ease;
            font-size: 16px;
            font-weight: 600;
        }
        button[data-baseweb="tab"]:hover { background-color: #dce0ea; color: black; }
        button[data-baseweb="tab"][aria-selected="true"] { background-color: #0366d6; color: white; }
        </style>
    """, unsafe_allow_html=True)
    
    # Cabe√ßalho bonito
    st.markdown("""
        <div style='display: flex; align-items: center; gap: 10px; margin-bottom: 20px;'>
            <img src='https://img.icons8.com/color/48/graph.png' width='40'/>
            <h1 style='display: inline; margin: 0; font-size: 2.4rem;'>Relat√≥rios Caixa Sangria</h1>
        </div>
    """, unsafe_allow_html=True)
    # ============ Helpers ============



    import unicodedata
    import re
    
    def _norm_txt(s: str) -> str:
        s = str(s or "").strip().lower()
        s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
        return s
    
    def eh_deposito_mask(df, cols_texto=None):
        """
        Retorna uma Series booleana marcando linhas de dep√≥sito.
        Crit√©rios por texto em colunas comuns (ajuste a lista se quiser).
        """
        if cols_texto is None:
            cols_texto = [
                "Descri√ß√£o Agrupada", "Descri√ß√£o", "Historico", "Hist√≥rico",
                "Categoria", "Obs", "Observa√ß√£o", "Tipo", "Tipo Movimento"
            ]
        cols_texto = [c for c in cols_texto if c in df.columns]
        if not cols_texto:
            return pd.Series(False, index=df.index)
    
        txt = df[cols_texto].astype(str).agg(" ".join, axis=1).map(_norm_txt)
    
        padrao = r"""
            \bdeposito\b        |   # 'deposito'/'dep√≥sito'
            \bdepsito\b         |   # varia√ß√µes sem acento
            \bdep\b             |   # abrevia√ß√£o comum
            credito\s+em\s+conta|
            transf(erencia)?\s*(p/?\s*banco|banco) |
            envio\s*para\s*banco|
            remessa\s* banco
        """
        return txt.str.contains(padrao, flags=re.IGNORECASE | re.VERBOSE, regex=True, na=False)
    
    def brl(v):
        try:
            return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        except Exception:
            return "R$ 0,00"

    def parse_valor_brl_sheets(x):
        """
        Normaliza valores vindos do Sheets para float (BRL):
        Aceita negativos '(...)' ou '-'. Remove 'R$', espa√ßos e pontos de milhar.
        Regras para quando n√£o h√° v√≠rgula: heur√≠sticas de casas decimais.
        """
        if isinstance(x, (int, float)):
            try:
                return float(x)
            except Exception:
                return 0.0

        s = str(x).strip()
        if s == "" or s.lower() in {"nan", "none"}:
            return 0.0

        neg = False
        if s.startswith("(") and s.endswith(")"):
            neg = True
            s = s[1:-1].strip()
        if s.startswith("-"):
            neg = True
            s = s[1:].strip()

        s = (s.replace("R$", "")
             .replace("\u00A0", "")
             .replace(" ", "")
             .replace(".", ""))

        if "," in s:
            inteiro, dec = s.rsplit(",", 1)
            inteiro = re.sub(r"\D", "", inteiro)
            dec     = re.sub(r"\D", "", dec)
            if dec == "":
                dec = "00"
            elif len(dec) == 1:
                dec = dec + "0"
            else:
                dec = dec[:2]
            num_str = f"{inteiro}.{dec}" if inteiro != "" else f"0.{dec}"
            try:
                val = float(num_str)
            except Exception:
                val = 0.0
        else:
            digits = re.sub(r"\D", "", s)
            if digits == "":
                val = 0.0
            else:
                n = len(digits)
                if n <= 3:
                    val = float(digits)
                elif n == 4:
                    if digits.endswith("00"):
                        val = float(digits) / 100.0
                    elif digits.endswith("0"):
                        val = float(digits) / 10.0
                    else:
                        val = float(digits)
                else:  # n >= 5
                    val = float(digits) / 100.0

        return -val if neg else val

    def _render_df(df, *, height=480):
        df = df.copy().reset_index(drop=True)
        seen, new_cols = {}, []
        for c in df.columns:
            s = "" if c is None else str(c)
            if s in seen:
                seen[s] += 1
                s = f"{s}_{seen[s]}"
            else:
                seen[s] = 0
            new_cols.append(s)
        df.columns = new_cols
        st.dataframe(df, use_container_width=True, height=height, hide_index=True)
        return df

    def pick_valor_col(cols):
        def norm(s):
            return re.sub(r"[\s\u00A0]+", " ", str(s)).strip().lower()
        nm = {c: norm(c) for c in cols}

        prefer = ["valor(r$)", "valor (r$)", "valor", "valor r$"]
        for want in prefer:
            for c, n in nm.items():
                if n == want:
                    return c

        for c, n in nm.items():
            if ("valor" in n
                and "valores" not in n
                and "google"  not in n
                and "sheet"   not in n):
                return c
        return None

    # ============ Carrega aba Sangria ============
    df_sangria = None
    try:
        ws_sangria = planilha_empresa.worksheet("Sangria")
        df_sangria = pd.DataFrame(ws_sangria.get_all_records())
        df_sangria.columns = [c.strip() for c in df_sangria.columns]
        if "Data" in df_sangria.columns:
            df_sangria["Data"] = pd.to_datetime(df_sangria["Data"], dayfirst=True, errors="coerce")
    except Exception as e:
        st.warning(f"‚ö†Ô∏è N√£o foi poss√≠vel carregar a aba 'Sangria': {e}")

# ============ Cabe√ßalho ============
#st.markdown("""
#<div style='display:flex;align-items:center;gap:10px;margin-bottom:12px;'>
#  <img src='https://img.icons8.com/color/48/cash-register.png' width='36'/>
#  <h1 style='margin:0;font-size:1.8rem;'>Relat√≥rios Caixa & Sangria</h1>
#</div>
#""", unsafe_allow_html=True)

# ============ Sub-Abas ============
sub_sangria, sub_caixa = st.tabs([
    "üí∏ Movimenta√ß√£o de Caixa",   # Anal√≠tico/Sint√©tico da Sangria
    "üß∞ Controle de Sangria"      # Comparativa Everest / Diferen√ßas
   
])

# -------------------------------
# Sub-aba: üí∏  Movimenta√ß√£o de Caixa (Anal√≠tico / Sint√©tico)
# -------------------------------
with sub_sangria:
    if df_sangria is None or df_sangria.empty:
        st.info("Sem dados de **sangria** dispon√≠veis.")
    else:
        from io import BytesIO

        # Base e colunas
        df = df_sangria.copy()
        df.columns = [str(c).strip() for c in df.columns]

        # Data obrigat√≥ria e normalizada
        if "Data" not in df.columns:
            st.error("A aba 'Sangria' precisa da coluna **Data**.")
            st.stop()
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)

        # Coluna de valor e parsing BRL -> float
        col_valor = pick_valor_col(df.columns)
        if not col_valor:
            st.error("N√£o encontrei a coluna de **valor** (ex.: 'Valor(R$)').")
            st.stop()
        df[col_valor] = df[col_valor].map(parse_valor_brl_sheets).astype(float)

        # Filtros
        c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.6, 1.6])
        with c1:
            dmin = pd.to_datetime(df["Data"].min(), errors="coerce")
            dmax = pd.to_datetime(df["Data"].max(), errors="coerce")
            today = pd.Timestamp.today().normalize()
            if pd.isna(dmin): dmin = today
            if pd.isna(dmax): dmax = today
            dt_inicio, dt_fim = st.date_input(
                "Per√≠odo",
                value=(dmax.date(), dmax.date()),
                min_value=dmin.date(),
                max_value=(dmax.date() if dmax >= dmin else dmin.date()),
                key="periodo_sangria_movi"
            )
        with c2:
            lojas = sorted(df.get("Loja", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
            lojas_sel = st.multiselect("Lojas", options=lojas, default=[], key="lojas_sangria_movi")
        with c3:
            descrs = sorted(df.get("Descri√ß√£o Agrupada", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
            descrs_sel = st.multiselect("Descri√ß√£o Agrupada", options=descrs, default=[], key="descr_sangria_movi")
        with c4:
            visao = st.selectbox(
                "Vis√£o do Relat√≥rio",
                options=["Anal√≠tico", "Sint√©tico"],
                index=0,
                key="visao_sangria_movi"
            )

        # Aplica filtros
        df_fil = df[(df["Data"].dt.date >= dt_inicio) & (df["Data"].dt.date <= dt_fim)].copy()
        if lojas_sel:
            df_fil = df_fil[df_fil["Loja"].astype(str).isin(lojas_sel)]
        if descrs_sel:
            df_fil = df_fil[df_fil["Descri√ß√£o Agrupada"].astype(str).isin(descrs_sel)]

        # Helper de formata√ß√£o BRL (apenas visual)
        def _fmt_brl_df(_df, col):
            _df[col] = _df[col].apply(
                lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                if isinstance(v, (int, float)) else v
            )
            return _df

        df_exibe = pd.DataFrame()

        # ====== Anal√≠tico ======
        if visao == "Anal√≠tico":
            grid = st.empty()

            df_base = df_fil.copy()
            df_base["Data"] = pd.to_datetime(df_base["Data"], errors="coerce").dt.normalize()
            df_base = df_base.sort_values(["Data"], na_position="last")

            total_val = df_base[col_valor].sum(min_count=1)
            total_row = {c: "" for c in df_base.columns}
            if "Loja" in total_row: total_row["Loja"] = "TOTAL"
            if "Data" in total_row: total_row["Data"] = pd.NaT
            if "Descri√ß√£o Agrupada" in total_row: total_row["Descri√ß√£o Agrupada"] = ""
            total_row[col_valor] = total_val

            df_exibe = pd.concat([pd.DataFrame([total_row]), df_base], ignore_index=True)

            # Datas (TOTAL vazio) e valor formatado
            df_exibe["Data"] = pd.to_datetime(df_exibe["Data"], errors="coerce").dt.strftime("%d/%m/%Y")
            df_exibe.loc[df_exibe.index == 0, "Data"] = ""
            df_exibe = _fmt_brl_df(df_exibe, col_valor)

            # Remove colunas t√©cnicas/ru√≠do
            aliases_remover = [
                "C√≥digo Everest", "Codigo Everest", "Cod Everest",
                "C√≥digo grupo Everest", "Codigo grupo Everest", "Cod Grupo Everest", "C√≥digo Grupo Everest",
                "M√™s", "Mes", "Ano", "Duplicidade", "Poss√≠vel Duplicidade", "Duplicado", "Sistema"
            ]
            df_exibe = df_exibe.drop(columns=[c for c in aliases_remover if c in df_exibe.columns], errors="ignore")

            grid.dataframe(df_exibe, use_container_width=True, hide_index=True)

            # Export Excel mantendo tipos
            df_export = pd.concat([pd.DataFrame([total_row]), df_base], ignore_index=True)
            df_export = df_export.drop(columns=[c for c in aliases_remover if c in df_export.columns], errors="ignore")
            df_export["Data"] = pd.to_datetime(df_export["Data"], errors="coerce")
            df_export[col_valor] = pd.to_numeric(df_export[col_valor], errors="coerce")

            buf = BytesIO()
            with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
                sh = "Anal√≠tico"
                df_export.to_excel(writer, sheet_name=sh, index=False)
                wb, ws = writer.book, writer.sheets[sh]
                header = wb.add_format({"bold": True,"align":"center","valign":"vcenter","bg_color":"#F2F2F2","border":1})
                date_f = wb.add_format({"num_format":"dd/mm/yyyy","border":1})
                money  = wb.add_format({"num_format":"R$ #,##0.00","border":1})
                text   = wb.add_format({"border":1})
                tot    = wb.add_format({"bold": True,"bg_color":"#FCE5CD","border":1})
                totm   = wb.add_format({"bold": True,"bg_color":"#FCE5CD","border":1,"num_format":"R$ #,##0.00"})

                for j, name in enumerate(df_export.columns):
                    ws.write(0, j, name, header)
                    width, fmt = 18, text
                    if name.lower() == "data": width, fmt = 12, date_f
                    if name == col_valor:      width, fmt = 16, money
                    if "loja"  in name.lower(): width = 28
                    if "grupo" in name.lower(): width = 22
                    ws.set_column(j, j, width, fmt)

                ws.set_row(1, None, tot)
                if pd.notna(df_export.iloc[0][col_valor]):
                    ws.write_number(1, list(df_export.columns).index(col_valor), float(df_export.iloc[0][col_valor]), totm)
                if "Loja" in df_export.columns:
                    ws.write_string(1, list(df_export.columns).index("Loja"), "TOTAL", tot)
                ws.freeze_panes(1, 0)

            buf.seek(0)
            st.download_button(
                label="‚¨áÔ∏è Baixar Excel",
                data=buf,  # ou buf.getvalue()
                file_name="Relatorio_Analitico_Sangria.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="dl_sangria_analitico"
            )


        # ====== Sint√©tico ======
        elif visao == "Sint√©tico":
            if "Loja" not in df_fil.columns:
                st.warning("Para 'Sint√©tico', preciso da coluna **Loja**.")
            else:
                tmp = df_fil.copy()
                tmp["Data"] = pd.to_datetime(tmp["Data"], errors="coerce").dt.normalize()

                # Garante 'Grupo'
                col_grupo = None
                for c in tmp.columns:
                    if str(c).strip().lower() == "grupo":
                        col_grupo = c; break
                if not col_grupo:
                    col_grupo = next((c for c in tmp.columns if "grupo" in str(c).lower() and "everest" not in str(c).lower()), None)
                if not col_grupo and "Loja" in tmp.columns:
                    mapa = df_empresa[["Loja", "Grupo"]].drop_duplicates()
                    tmp = tmp.merge(mapa, on="Loja", how="left")
                    col_grupo = "Grupo"

                group_cols = [c for c in [col_grupo, "Loja", "Data"] if c]
                df_agg = tmp.groupby(group_cols, as_index=False)[col_valor].sum()

                ren = {col_valor: "Sangria"}
                if col_grupo and col_grupo != "Grupo":
                    ren[col_grupo] = "Grupo"
                df_agg = df_agg.rename(columns=ren).sort_values(["Data", "Grupo", "Loja"], na_position="last")

                total_sangria = df_agg["Sangria"].sum(min_count=1)
                linha_total = pd.DataFrame({"Grupo":["TOTAL"], "Loja":[""], "Data":[pd.NaT], "Sangria":[total_sangria]})
                df_exibe = pd.concat([linha_total, df_agg], ignore_index=True)

                # Exibi√ß√£o
                df_show = df_exibe.copy()
                df_show["Data"] = pd.to_datetime(df_show["Data"], errors="coerce").dt.strftime("%d/%m/%Y").fillna("")
                df_show["Sangria"] = df_show["Sangria"].apply(lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

                st.dataframe(df_show[["Grupo","Loja","Data","Sangria"]], use_container_width=True, hide_index=True)

                # Export Excel
                df_exp = df_exibe[["Grupo","Loja","Data","Sangria"]].copy()
                df_exp["Data"] = pd.to_datetime(df_exp["Data"], errors="coerce")
                df_exp["Sangria"] = pd.to_numeric(df_exp["Sangria"], errors="coerce")

                buf = BytesIO()
                with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
                    sh = "Sint√©tico"
                    df_exp.to_excel(writer, sheet_name=sh, index=False)
                    wb, ws = writer.book, writer.sheets[sh]
                    header = wb.add_format({"bold": True,"align":"center","valign":"vcenter","bg_color":"#F2F2F2","border":1})
                    date_f = wb.add_format({"num_format":"dd/mm/yyyy","border":1})
                    money  = wb.add_format({"num_format":"R$ #,##0.00","border":1})
                    text   = wb.add_format({"border":1})
                    tot    = wb.add_format({"bold": True,"bg_color":"#FCE5CD","border":1})
                    totm   = wb.add_format({"bold": True,"bg_color":"#FCE5CD","border":1,"num_format":"R$ #,##0.00"})

                    for j, name in enumerate(["Grupo","Loja","Data","Sangria"]):
                        ws.write(0, j, name, header)
                    ws.set_column("A:A", 20, text)
                    ws.set_column("B:B", 28, text)
                    ws.set_column("C:C", 12, date_f)
                    ws.set_column("D:D", 14, money)

                    ws.set_row(1, None, tot)
                    if pd.notna(df_exp.iloc[0]["Sangria"]):
                        ws.write_number(1, 3, float(df_exp.iloc[0]["Sangria"]), totm)
                    ws.write_string(1, 0, "TOTAL", tot)
                    ws.freeze_panes(1, 0)

                buf.seek(0)
                st.download_button(
                    label="‚¨áÔ∏è Baixar Excel",
                    data=buf,
                    file_name="Relatorio_Sintetico_Sangria.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key="dl_sangria_sintetico"
                )

# -------------------------------
# Sub-aba: üß∞ CONTROLE DE SANGRIA (Comparativa Everest / Diferen√ßas)
# -------------------------------
with sub_caixa:
    if df_sangria is None or df_sangria.empty:
        st.info("Sem dados de **sangria** dispon√≠veis.")
    else:
        from io import BytesIO
        import unicodedata, re, os
        import pandas as pd
        from datetime import datetime

        # ===== helpers =====
        def _norm_txt(s: str) -> str:
            s = str(s or "").strip().lower()
            s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("utf-8")
            return s

        def eh_deposito_mask(df, cols_texto=None):
            if cols_texto is None:
                cols_texto = [
                    "Descri√ß√£o Agrupada","Descri√ß√£o","Historico","Hist√≥rico",
                    "Categoria","Obs","Observa√ß√£o","Tipo","Tipo Movimento"
                ]
            cols_texto = [c for c in cols_texto if c in df.columns]
            if not cols_texto:
                return pd.Series(False, index=df.index)
            txt = df[cols_texto].astype(str).agg(" ".join, axis=1).map(_norm_txt)
            padrao = r"""
                \bdeposito\b | \bdepsito\b | \bdep\b |
                credito\s+em\s+conta | envio\s*para\s*banco |
                transf(erencia)?\s*(p/?\s*banco|banco)
            """
            return txt.str.contains(padrao, flags=re.IGNORECASE | re.VERBOSE, regex=True, na=False)

        def brl(v):
            try:
                return f"R$ {float(v):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            except Exception:
                return "R$ 0,00"

        # ===== base =====
        df = df_sangria.copy()
        df.columns = [str(c).strip() for c in df.columns]

        if "Data" not in df.columns:
            st.error("A aba 'Sangria' precisa da coluna **Data**.")
            st.stop()
        df["Data"] = pd.to_datetime(df["Data"], errors="coerce", dayfirst=True)

        col_valor = pick_valor_col(df.columns)
        if not col_valor:
            st.error("N√£o encontrei a coluna de **valor** (ex.: 'Valor(R$)').")
            st.stop()
        df[col_valor] = df[col_valor].map(parse_valor_brl_sheets).astype(float)

        # Filtros
        c1, c2, c3, c4 = st.columns([1.2, 1.2, 1.6, 1.6])
        with c1:
            dmin = pd.to_datetime(df["Data"].min(), errors="coerce")
            dmax = pd.to_datetime(df["Data"].max(), errors="coerce")
            today = pd.Timestamp.today().normalize()
            if pd.isna(dmin): dmin = today
            if pd.isna(dmax): dmax = today
            dt_inicio, dt_fim = st.date_input(
                "Per√≠odo",
                value=(dmax.date(), dmax.date()),
                min_value=dmin.date(),
                max_value=(dmax.date() if dmax >= dmin else dmin.date()),
                key="caixa_periodo_cmp",
            )
        with c2:
            lojas = sorted(df.get("Loja", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
            lojas_sel = st.multiselect("Lojas", options=lojas, default=[], key="caixa_lojas_cmp")
        with c3:
            descrs = sorted(df.get("Descri√ß√£o Agrupada", pd.Series(dtype=str)).dropna().astype(str).unique().tolist())
            descrs_sel = st.multiselect("Descri√ß√£o Agrupada", options=descrs, default=[], key="caixa_descr_cmp")
        with c4:
            visao = st.selectbox(
                "Vis√£o do Relat√≥rio",
                options=["Comparativa Everest"],  # foquei na comparativa
                index=0,
                key="caixa_visao_cmp",
            )

        # aplica filtros
        df_fil = df[(df["Data"].dt.date >= dt_inicio) & (df["Data"].dt.date <= dt_fim)].copy()
        if lojas_sel:
            df_fil = df_fil[df_fil["Loja"].astype(str).isin(lojas_sel)]
        if descrs_sel:
            df_fil = df_fil[df_fil["Descri√ß√£o Agrupada"].astype(str).isin(descrs_sel)]

        df_exibe = pd.DataFrame()

        # ======= Comparativa =======
        if visao == "Comparativa Everest":
            base = df_fil.copy()

            if "Data" not in base.columns or "C√≥digo Everest" not in base.columns or not col_valor:
                st.error("‚ùå Preciso de 'Data', 'C√≥digo Everest' e coluna de valor na aba Sangria.")
            else:
                # normaliza√ß√£o
                base["Data"] = pd.to_datetime(base["Data"], dayfirst=True, errors="coerce").dt.normalize()
                base[col_valor] = pd.to_numeric(base[col_valor], errors="coerce").fillna(0.0)
                base["C√≥digo Everest"] = base["C√≥digo Everest"].astype(str).str.extract(r"(\d+)")

                # --- EXCLUI DEP√ìSITOS (somente lado Sistema/Colibri) ---
                mask_dep_sys = eh_deposito_mask(base)
                with st.expander("üîé Ver dep√≥sitos removidos (Colibri/CISS)"):
                    audit = base.loc[mask_dep_sys, :].copy()
                    if col_valor in audit.columns:
                        audit[col_valor] = audit[col_valor].map(brl)
                    st.dataframe(audit, use_container_width=True, hide_index=True)

                base = base.loc[~mask_dep_sys].copy()

                # agrega Sistema (j√° sem dep√≥sitos)
                df_sys = (
                    base.groupby(["C√≥digo Everest","Data"], as_index=False)[col_valor]
                        .sum()
                        .rename(columns={col_valor:"Sangria (Colibri/CISS)"})
                )

                # --- Everest ---
                ws_ev = planilha_empresa.worksheet("Sangria Everest")
                df_ev = pd.DataFrame(ws_ev.get_all_records())
                df_ev.columns = [c.strip() for c in df_ev.columns]

                def _norm(s): return re.sub(r"[^a-z0-9]", "", str(s).lower())
                cmap = {_norm(c): c for c in df_ev.columns}
                col_emp   = cmap.get("empresa")
                col_dt_ev = next((orig for norm, orig in cmap.items()
                                  if norm in ("dlancamento","dlancament","dlanamento","datadelancamento","data")), None)
                col_val_ev= next((orig for norm, orig in cmap.items()
                                  if norm in ("valorlancamento","valorlancament","valorlcto","valor")), None)
                col_fant  = next((orig for norm, orig in cmap.items()
                                  if norm in ("fantasiaempresa","fantasia")), None)

                if not all([col_emp, col_dt_ev, col_val_ev]):
                    st.error("‚ùå Na 'Sangria Everest' preciso de 'Empresa', 'D. Lan√ßamento' e 'Valor Lancamento'.")
                else:
                    de = df_ev.copy()
                    de["C√≥digo Everest"]   = de[col_emp].astype(str).str.extract(r"(\d+)")
                    de["Fantasia Everest"] = de[col_fant] if col_fant else ""
                    de["Data"]             = pd.to_datetime(de[col_dt_ev], dayfirst=True, errors="coerce").dt.normalize()
                    de["Valor Lancamento"] = de[col_val_ev].map(parse_valor_brl_sheets).astype(float)
                    de = de[(de["Data"].dt.date >= dt_inicio) & (de["Data"].dt.date <= dt_fim)]
                    de["Sangria Everest"]  = de["Valor Lancamento"].abs()

                    def _pick_first(s):
                        s = s.dropna().astype(str).str.strip()
                        s = s[s != ""]
                        return s.iloc[0] if not s.empty else ""
                    de_agg = (
                        de.groupby(["C√≥digo Everest","Data"], as_index=False)
                          .agg({"Sangria Everest":"sum","Fantasia Everest": _pick_first})
                    )

                    cmp = df_sys.merge(de_agg, on=["C√≥digo Everest","Data"], how="outer", indicator=True)
                    cmp["Sangria (Colibri/CISS)"] = cmp["Sangria (Colibri/CISS)"].fillna(0.0)
                    cmp["Sangria Everest"]        = cmp["Sangria Everest"].fillna(0.0)

                    # mapeamento Loja/Grupo
                    mapa = df_empresa.copy()
                    mapa.columns = [str(c).strip() for c in mapa.columns]
                    if "C√≥digo Everest" in mapa.columns:
                        mapa["C√≥digo Everest"] = mapa["C√≥digo Everest"].astype(str).str.extract(r"(\d+)")
                        cmp = cmp.merge(mapa[["C√≥digo Everest","Loja","Grupo"]].drop_duplicates(),
                                        on="C√≥digo Everest", how="left")

                    # fallback LOJA = Fantasia (linhas apenas do Everest)
                    cmp["Loja"] = cmp["Loja"].astype(str)
                    so_everest = (cmp["_merge"] == "right_only") & (cmp["Loja"].isin(["", "nan"]))
                    cmp.loc[so_everest, "Loja"] = cmp.loc[so_everest, "Fantasia Everest"]
                    cmp["Nao Mapeada?"] = so_everest

                    cmp["Diferen√ßa"] = cmp["Sangria (Colibri/CISS)"] - cmp["Sangria Everest"]

                    cmp = cmp[["Grupo","Loja","C√≥digo Everest","Data",
                               "Sangria (Colibri/CISS)","Sangria Everest","Diferen√ßa","Nao Mapeada?"]
                             ].sort_values(["Grupo","Loja","C√≥digo Everest","Data"])

                    total = {
                        "Grupo":"TOTAL","Loja":"","C√≥digo Everest":"","Data":pd.NaT,
                        "Sangria (Colibri/CISS)": cmp["Sangria (Colibri/CISS)"].sum(),
                        "Sangria Everest":        cmp["Sangria Everest"].sum(),
                        "Diferen√ßa":              cmp["Diferen√ßa"].sum(),
                        "Nao Mapeada?": False
                    }
                    df_exibe = pd.concat([pd.DataFrame([total]), cmp], ignore_index=True)

                    # ---- render no app
                    df_show = df_exibe.copy()
                    df_show["Data"] = pd.to_datetime(df_show["Data"], errors="coerce").dt.strftime("%d/%m/%Y").fillna("")
                    for c in ["Sangria (Colibri/CISS)","Sangria Everest","Diferen√ßa"]:
                        df_show[c] = df_show[c].apply(
                            lambda v: f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X",".")
                            if isinstance(v,(int,float)) else v
                        )

                    if "Nao Mapeada?" in df_show.columns and "Loja" in df_show.columns:
                        view = df_show.drop(columns=["Nao Mapeada?"], errors="ignore").copy()
                        mask_nm = (
                            df_show["Nao Mapeada?"].astype(bool)
                            if "Nao Mapeada?" in df_show.columns
                            else pd.Series(False, index=df_show.index)
                        )
                        def _paint_row(row: pd.Series):
                            styles = [""] * len(row.index)
                            try:
                                if mask_nm.loc[row.name] and "Loja" in row.index:
                                    idx = list(row.index).index("Loja")
                                    styles[idx] = "color: red; font-weight: 700"
                            except Exception:
                                pass
                            return styles
                        st.dataframe(view.style.apply(_paint_row, axis=1), use_container_width=True, height=520)
                    else:
                        st.dataframe(df_show.drop(columns=["Nao Mapeada?"], errors="ignore"),
                                     use_container_width=True, height=520)

                    # ========= EXPORTA√á√ÉO (com slicers quando poss√≠vel) =========
                    # ======= EXPORTA√á√ÉO COM SLICERS (Excel Desktop + XlsxWriter >= 3.2.0) =======
                    # ========= EXPORTA√á√ÉO VIA TEMPLATE (mant√©m segmenta√ß√µes) =========
                    from io import BytesIO
                    import os, re
                    import pandas as pd
                    import streamlit as st
                    from openpyxl import load_workbook
                    from openpyxl.utils import get_column_letter, column_index_from_string
                    
                    TEMPLATE_NOME = "modelo_segmentacao_sangria.xlsx"  # coloque no mesmo diret√≥rio do .py
                    
                    def _localizar_template(nome: str) -> str | None:
                        candidatos = [
                            os.path.join(os.getcwd(), nome),
                            os.path.join(os.path.dirname(__file__), nome) if "__file__" in globals() else None,
                            nome,
                        ]
                        for p in candidatos:
                            if p and os.path.exists(p):
                                return p
                        return None
                    
                    def _normalizar_campo_valor(df: pd.DataFrame):
                        # renomeia caso venha "Sangria (Sistema)" do seu fluxo
                        if "Sangria (Sistema)" in df.columns and "Sangria (Colibri/CISS)" not in df.columns:
                            df = df.rename(columns={"Sangria (Sistema)": "Sangria (Colibri/CISS)"})
                        return df
                    
                    def _preparar_df_para_template(cmp: pd.DataFrame, headers_template: list[str]) -> pd.DataFrame:
                        df = cmp.copy()
                        df = df.drop(columns=["Nao Mapeada?"], errors="ignore")
                        df = _normalizar_campo_valor(df)
                    
                        # Base temporal
                        df["Data"] = pd.to_datetime(df["Data"], errors="coerce").dt.normalize()
                        df["Ano"]  = df["Data"].dt.year
                        df["M√™s"]  = df["Data"].dt.month
                    
                        # Num√©ricos
                        for c in ["Sangria (Colibri/CISS)", "Sangria Everest", "Diferen√ßa"]:
                            if c in df.columns:
                                df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0.0)
                    
                        # Ajuste do cabe√ßalho "M√™s" x "Mes" conforme o template
                        headers = headers_template[:]  # c√≥pia
                        if "Mes" in headers and "M√™s" in df.columns:
                            df = df.rename(columns={"M√™s": "Mes"})
                        elif "M√™s" in headers and "Mes" in df.columns:
                            df = df.rename(columns={"Mes": "M√™s"})
                    
                        # Reordena e mant√©m s√≥ o que o template pede
                        keep = [h for h in headers if h in df.columns]
                        df = df[keep].copy()
                        return df
                    
                    def _ler_headers_tabela(ws, nome_tabela: str) -> list[str]:
                        tbl = ws.tables.get(nome_tabela, None)
                        if tbl is None:
                            raise RuntimeError(f"O template n√£o tem a Tabela '{nome_tabela}' na planilha.")
                        ref_ini, ref_fim = tbl.ref.split(":")
                        m = re.match(r"([A-Z]+)(\d+)", ref_ini)
                        col_ini_letters, row_ini = m.group(1), int(m.group(2))
                        col_ini = column_index_from_string(col_ini_letters)
                    
                        # L√™ a linha do cabe√ßalho da tabela (row_ini)
                        m2 = re.match(r"[A-Z]+(\d+)", ref_fim)
                        col_fim_letters = re.match(r"([A-Z]+)\d+", ref_fim).group(1)
                        col_fim = column_index_from_string(col_fim_letters)
                    
                        headers = []
                        for j in range(col_ini, col_fim + 1):
                            headers.append(ws.cell(row=row_ini, column=j).value)
                        return headers
                    
                    def exportar_via_template(cmp: pd.DataFrame, caminho_template: str,
                                              planilha="Dados", nome_tabela="tbl_dados") -> BytesIO:
                        wb = load_workbook(caminho_template)
                        if planilha not in wb.sheetnames:
                            raise RuntimeError(f"O template n√£o possui a planilha '{planilha}'.")
                        ws = wb[planilha]
                    
                        # L√™ headers REAIS da tabela no template
                        tbl = ws.tables.get(nome_tabela, None)
                        if tbl is None:
                            raise RuntimeError(f"O template n√£o tem a Tabela '{nome_tabela}' na planilha '{planilha}'.")
                        headers_template = _ler_headers_tabela(ws, nome_tabela)
                    
                        # Descobre in√≠cio da tabela
                        ref_ini, ref_fim = tbl.ref.split(":")
                        m = re.match(r"([A-Z]+)(\d+)", ref_ini)
                        col_ini_letters, row_ini = m.group(1), int(m.group(2))
                        col_ini = column_index_from_string(col_ini_letters)
                    
                        # Prepara DF nos headers do template
                        df = _preparar_df_para_template(cmp, headers_template)
                    
                        # 1) sobrescreve cabe√ßalhos (garante match exato com tabela/slicers)
                        for j, col in enumerate(headers_template, start=col_ini):
                            ws.cell(row=row_ini, column=j, value=col)
                    
                        # 2) apaga linhas antigas abaixo do header (mant√©m s√≥ o cabe√ßalho)
                        old_end_row = int(re.match(r"[A-Z]+(\d+)", ref_fim).group(1))
                        if old_end_row > row_ini:
                            ws.delete_rows(row_ini + 1, old_end_row - row_ini)
                    
                        # 3) escreve dados
                        date_fmt  = "dd/mm/yyyy"
                        money_fmt = '"R$" #,##0.00'
                        int_fmt   = "0"
                    
                        for i, (_, row) in enumerate(df.iterrows(), start=row_ini + 1):
                            for j, col in enumerate(headers_template, start=col_ini):
                                cell = ws.cell(row=i, column=j)
                                val  = row[col] if col in row.index else None
                                if col == "Data" and pd.notna(val):
                                    cell.value = pd.to_datetime(val).date()
                                    cell.number_format = date_fmt
                                elif col in ("Ano", "M√™s", "Mes", "C√≥digo Everest"):
                                    cell.value = int(val) if pd.notna(val) else None
                                    cell.number_format = int_fmt
                                elif col in ("Sangria (Colibri/CISS)", "Sangria Everest", "Diferen√ßa"):
                                    cell.value = float(val) if pd.notna(val) else None
                                    cell.number_format = money_fmt
                                else:
                                    cell.value = "" if pd.isna(val) else val
                    
                        # 4) atualiza o range da tabela para cobrir todas as linhas/colunas
                        end_row = row_ini + len(df)
                        end_col = col_ini + len(headers_template) - 1
                        new_ref = f"{get_column_letter(col_ini)}{row_ini}:{get_column_letter(end_col)}{end_row}"
                        tbl.ref = new_ref
                    
                        # 5) salva em mem√≥ria
                        out = BytesIO()
                        wb.save(out)
                        out.seek(0)
                        return out
                    
                    # ===== uso =====
                    caminho_tpl = _localizar_template(TEMPLATE_NOME)
                    if not caminho_tpl:
                        st.error(
                            f"Template **{TEMPLATE_NOME}** n√£o encontrado.\n"
                            f"Coloque-o no mesmo diret√≥rio do arquivo .py.\n\n"
                            f"Requisitos do template:\n"
                            f"‚Ä¢ Planilha: 'Dados'\n"
                            f"‚Ä¢ Tabela:   'tbl_dados' iniciando em A1 (cabe√ßalho na linha da tabela)\n"
                            f"‚Ä¢ Segmenta√ß√µes (inseridas no Excel Desktop) para: Ano, M√™s/Mes, Grupo, Loja\n"
                            f"‚Ä¢ Cabe√ßalhos da tabela exatamente como est√£o nas segmenta√ß√µes."
                        )
                    else:
                        try:
                            arquivo = exportar_via_template(cmp, caminho_tpl)
                            st.download_button(
                                "‚¨áÔ∏è Baixar Excel",
                                data=arquivo,
                                file_name="Sangria_Controle.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                key="dl_sangria_controle_excel"
                            )
                            st.caption(f"Template usado: {os.path.basename(caminho_tpl)} ‚Äî segmenta√ß√µes preservadas.")
                        except Exception as e:
                            st.error(f"Falha ao preencher o template: {type(e).__name__}: {e}")
